# -*- coding: utf-8 -*-
"""Stream-filter the national business-registry dump (853 city xlsx files) for
debt-collection-industry firms, and tally all-firm entry/exit counts per
city-month in the same pass.

Outputs (<restricted-source-path>):
  hits\\{file}.csv  matched firms: name, status, founding/approval dates, scope
                    snippet, matched keyword+field, geography
  agg\\{file}.csv   per (city, month): all-firm foundings; deregistered-firm
                    approval-month counts (exit clock)
Checkpointed per input file; safe to rerun. Keywords kept tight
(催收|讨债|商账|收数); 收数 false-hits (回收数码/数控/数据) are flagged via the
stored snippet for the audit pass, not silently dropped.
"""

# Replication-package paths
from pathlib import Path as _ReplicationPath
import os as _ReplicationOS
_REP_PROJECT = _ReplicationPath(__file__).resolve().parents[1]
_REP_PACKAGE = _REP_PROJECT.parent
_REP_RESTRICTED = _REP_PACKAGE / 'restricted_data'
_REP_JUDGMENTS = _ReplicationPath(_ReplicationOS.environ.get(
    'HWIH_JUDGMENT_ARCHIVE', _REP_RESTRICTED / 'judgment_archive'))
_REP_CASE_ARCHIVE = _ReplicationPath(_ReplicationOS.environ.get(
    'HWIH_CASE_LEVEL_ARCHIVE', _REP_RESTRICTED / 'case_level_archive.parquet'))
_REP_MORTALITY = _ReplicationPath(_ReplicationOS.environ.get(
    'HWIH_CDC_SOURCE_ROOT', _REP_RESTRICTED / 'mortality_volumes'))
_REP_REGISTRY = _ReplicationPath(_ReplicationOS.environ.get(
    'HWIH_REGISTRY_ROOT', _REP_RESTRICTED / 'firm_registry'))
_REP_BAIDU = _ReplicationPath(_ReplicationOS.environ.get(
    'HWIH_BAIDU_ROOT', _REP_RESTRICTED / 'baidu'))
_REP_INTERVIEWS = _ReplicationPath(_ReplicationOS.environ.get(
    'HWIH_INTERVIEWS_ROOT', _REP_RESTRICTED / 'interviews'))
import csv, os, re, sys, glob
from multiprocessing import Pool

SRC = glob.glob(str(_REP_PACKAGE / "restricted_data" / "source_data").replace('\\', '/'))[0]
OUT = str(_REP_REGISTRY)
KW = re.compile(r"催收|讨债|商账|收数")

COLS = ["企业名称", "经营状态", "成立日期", "核准日期", "经营范围",
        "所属省份", "所属城市", "所属区县"]


def ym(v):
    s = str(v)[:7]
    return s if re.match(r"^\d{4}-\d{2}", s) else ""


def process(fp):
    base = os.path.splitext(os.path.basename(fp))[0]
    hfp = os.path.join(OUT, "hits", base + ".csv")
    afp = os.path.join(OUT, "agg", base + ".csv")
    if os.path.exists(afp):
        return base, -1, -1
    try:
        from openpyxl import load_workbook
        wb = load_workbook(fp, read_only=True, data_only=True)
        ws = wb.worksheets[0]
        rows = ws.iter_rows(values_only=True)
        header = [str(c).strip() if c is not None else "" for c in next(rows)]
        ix = {}
        for c in COLS:
            ix[c] = header.index(c) if c in header else None
        entry, exit_ = {}, {}
        nhit = ntot = 0
        os.makedirs(os.path.dirname(hfp), exist_ok=True)
        os.makedirs(os.path.dirname(afp), exist_ok=True)
        with open(hfp + ".tmp", "w", newline="", encoding="utf-8-sig") as hf:
            w = csv.writer(hf)
            w.writerow(COLS[:4] + ["matched_kw", "matched_field", "snippet"]
                       + COLS[5:])
            for r in rows:
                ntot += 1
                def g(c):
                    i = ix[c]
                    return ("" if i is None or i >= len(r) or r[i] is None
                            else str(r[i]))
                city = g("所属城市") or g("所属省份")
                em, xm = ym(g("成立日期")), ""
                if em:
                    entry[(city, em)] = entry.get((city, em), 0) + 1
                st = g("经营状态")
                if "注销" in st or "吊销" in st:
                    xm = ym(g("核准日期"))
                    if xm:
                        exit_[(city, xm)] = exit_.get((city, xm), 0) + 1
                name, scope = g("企业名称"), g("经营范围")
                m = KW.search(name)
                field = "名称" if m else ""
                if not m:
                    m = KW.search(scope)
                    field = "经营范围" if m else ""
                if m:
                    nhit += 1
                    src = name if field == "名称" else scope
                    a = max(0, m.start() - 25)
                    snip = src[a:m.end() + 25].replace("\n", " ")
                    w.writerow([name, st, g("成立日期"), g("核准日期"),
                                m.group(0), field, snip, g("所属省份"),
                                g("所属城市"), g("所属区县")])
        wb.close()
        with open(afp + ".tmp", "w", newline="", encoding="utf-8-sig") as af:
            w = csv.writer(af)
            w.writerow(["city", "month", "entries_all", "exits_all"])
            for k in sorted(set(entry) | set(exit_)):
                w.writerow([k[0], k[1], entry.get(k, 0), exit_.get(k, 0)])
        os.replace(hfp + ".tmp", hfp)
        os.replace(afp + ".tmp", afp)
        return base, ntot, nhit
    except Exception as e:
        return base, -2, f"{type(e).__name__}: {e}"


if __name__ == "__main__":
    sys.stdout.reconfigure(encoding="utf-8")
    files = [f for f in glob.glob(os.path.join(SRC, "*.xlsx"))
             if ".baiduyun" not in f]
    print(f"{len(files)} files", flush=True)
    done = fail = 0
    with Pool(6) as pool:
        for base, ntot, nhit in pool.imap_unordered(process, files):
            if ntot == -1:
                continue
            if ntot == -2:
                fail += 1
                print(f"ERR {base}: {nhit}", flush=True)
                continue
            done += 1
            print(f"ok {base}: rows={ntot} hits={nhit}", flush=True)
    print(f"DONE processed={done} failed={fail}", flush=True)
